"""Excel export helpers for the budget tree and the application list.

The module imports `openpyxl` lazily, only on the export path, so an environment without
the package still loads. An endpoint passes in data that it filtered already. This module
knows no database. It turns rows into bytes.
"""

from __future__ import annotations

import json
from collections.abc import Iterable, Mapping, Sequence
from datetime import datetime
from decimal import Decimal
from io import BytesIO
from typing import TYPE_CHECKING, Any

if TYPE_CHECKING:  # pragma: no cover - types only
    from app.modules.applications.schemas import ApplicationListItem
    from app.modules.budget.tree_schemas import BudgetTreeNodeOut

XLSX_MEDIA_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


def _num(value: Decimal | float | None) -> float | None:
    return float(value) if value is not None else None


# Leading characters that can start an active formula or a DDE call when Excel or
# LibreOffice opens the file. This is CSV and XLSX formula injection.
_FORMULA_PREFIXES: tuple[str, ...] = ("=", "+", "-", "@", "\t", "\r")


def _safe(value: Any) -> Any:
    """Neutralize formula injection.

    A string that starts with a dangerous character gets a leading apostrophe. The
    dangerous characters are `=`, `+`, `-`, `@`, tab and CR. A spreadsheet then treats
    the value as text and not as a formula. A value that is not a string, such as a
    number or `None`, passes through unchanged.
    """
    if isinstance(value, str) and value[:1] in _FORMULA_PREFIXES:
        return "'" + value
    return value


def _append(worksheet: Any, row: Sequence[Any]) -> None:
    """Append a row with `ws.append` and protect every cell against formula injection."""
    worksheet.append([_safe(cell) for cell in row])


def _autosize(worksheet: Any, headers: Sequence[str]) -> None:
    """Fit each column width roughly to its longest cell."""
    widths = [len(str(h)) for h in headers]
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        for i, cell in enumerate(row):
            widths[i] = max(widths[i], len(str(cell)) if cell is not None else 0)
    from openpyxl.utils import get_column_letter

    for i, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(i)].width = min(width + 2, 60)


def _header_row(worksheet: Any, headers: Sequence[str]) -> None:
    from openpyxl.styles import Font

    worksheet.append(list(headers))
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
    worksheet.freeze_panes = "A2"


def _iter_nodes(
    nodes: Iterable[BudgetTreeNodeOut], depth: int = 0
) -> Iterable[tuple[int, BudgetTreeNodeOut]]:
    for node in nodes:
        yield depth, node
        yield from _iter_nodes(node.children, depth + 1)


def build_budget_workbook(
    roots: Sequence[BudgetTreeNodeOut],
    *,
    fiscal_year_labels: dict[Any, str],
    fiscal_year_id: Any | None = None,
) -> bytes:
    """Build the budget tree as `.xlsx` bytes, with one sheet per fiscal year.

    Args:
        roots: The tree, already reduced to the visible selection of Gremium or
            subtree.
        fiscal_year_labels: Maps a fiscal year id to the sheet title. A year without an
            entry gets the sheet title `HHJ`.
        fiscal_year_id: Filter to a single fiscal year and write only its sheet. Pass
            `None` to write a sheet for every fiscal year.
    """
    from openpyxl import Workbook

    headers = [
        "Kostenstelle",
        "Schlüssel",
        "Zugeteilt",
        "Gebunden",
        "Beantragt",
        "Verfügbar",
        "Währung",
    ]
    nodes = list(_iter_nodes(roots))

    # Collect the fiscal years in tree order, which keeps the sheet order stable.
    fy_order: list[Any] = []
    for _depth, node in nodes:
        for alloc in node.by_fiscal_year:
            if fiscal_year_id is not None and alloc.fiscal_year_id != fiscal_year_id:
                continue
            if alloc.fiscal_year_id not in fy_order:
                fy_order.append(alloc.fiscal_year_id)

    wb = Workbook()
    default_ws = wb.active
    if default_ws is not None:
        wb.remove(default_ws)

    if not fy_order:
        ws = wb.create_sheet(title="Budget")
        _header_row(ws, headers)
        _autosize(ws, headers)
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    used_titles: set[str] = set()
    for fy in fy_order:
        label = fiscal_year_labels.get(fy, "HHJ")
        ws = wb.create_sheet(title=_sheet_title(label, used_titles))
        _header_row(ws, headers)
        for depth, node in nodes:
            indented = ("    " * depth) + node.name
            alloc = next(
                (a for a in node.by_fiscal_year if a.fiscal_year_id == fy), None
            )
            if alloc is None:
                _append(ws, [indented, node.path_key, None, None, None, None, node.currency])
            else:
                _append(
                    ws,
                    [
                        indented,
                        node.path_key,
                        _num(alloc.allocated),
                        _num(alloc.committed),
                        _num(alloc.requested),
                        _num(alloc.available),
                        node.currency,
                    ]
                )
        _autosize(ws, headers)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _sheet_title(label: str, used: set[str]) -> str:
    r"""Build a unique sheet name that Excel accepts.

    The name holds 31 characters at most and none of the characters `[]:*?/\`.
    """
    safe = "".join("_" if c in '[]:*?/\\' else c for c in label).strip() or "HHJ"
    safe = safe[:31]
    base = safe
    n = 2
    while safe in used:
        suffix = f" ({n})"
        safe = base[: 31 - len(suffix)] + suffix
        n += 1
    used.add(safe)
    return safe


def build_applications_workbook(
    items: Sequence[ApplicationListItem],
    *,
    type_names: dict[Any, str],
    gremium_names: dict[Any, str],
    locale: str = "de",
) -> bytes:
    """Build the application list as `.xlsx` bytes, in the order and filter as passed."""
    from openpyxl import Workbook

    headers = [
        "Titel",
        "Antragstyp",
        "Status",
        "Gremium",
        "Betrag",
        "Währung",
        "Erstellt",
        "Aktualisiert",
    ]
    wb = Workbook()
    ws = wb.active
    assert ws is not None  # noqa: S101 - openpyxl always returns an active sheet
    ws.title = "Anträge"
    _header_row(ws, headers)

    for item in items:
        state_label = ""
        if item.state is not None:
            label = item.state.label or {}
            state_label = label.get(locale) or label.get("de") or label.get("en") or ""
        _append(
            ws,
            [
                item.title or "",
                type_names.get(item.type_id, ""),
                state_label,
                gremium_names.get(item.gremium_id, "") if item.gremium_id else "",
                _num(item.amount),
                item.currency or "",
                _fmt_dt(item.created_at),
                _fmt_dt(item.updated_at),
            ],
        )

    _autosize(ws, headers)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_expenses_workbook(items: Iterable[Any], locale: str = "de") -> bytes:
    """Build the bookings, expenses and income, as `.xlsx` bytes.

    The rows keep the order and the filter as passed.
    """
    from openpyxl import Workbook

    kind_label = (
        {"expense": "Ausgabe", "income": "Einnahme"}
        if locale == "de"
        else {"expense": "Expense", "income": "Income"}
    )
    headers = [
        "Datum", "Art", "Beschreibung", "Kostenstelle", "Antrag", "Konto", "Betrag", "Währung",
    ]
    wb = Workbook()
    ws = wb.active
    assert ws is not None  # noqa: S101 - openpyxl always returns an active sheet
    ws.title = "Buchungen"
    _header_row(ws, headers)
    for e in items:
        _append(
            ws,
            [
                _fmt_dt(e.created_at),
                kind_label.get(e.kind, e.kind),
                e.description or "",
                e.path_key or "",
                e.application_title or "",
                e.account_name or "",
                _num(e.amount),
                e.currency or "",
            ],
        )
    _autosize(ws, headers)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_auskunft_workbook(
    *,
    email: str,
    applications: Sequence[Mapping[str, Any]],
    versions: Sequence[Mapping[str, Any]],
    principal: Mapping[str, Any] | None,
) -> bytes:
    """Build the GDPR subject-access export (Art. 15) as `.xlsx` bytes.

    The workbook holds every personal data record stored for `email`. The function is
    database-agnostic and takes already-prepared rows. It writes three sheets: the
    account of the principal, the applications with `data` as JSON, and the versions.
    """
    from openpyxl import Workbook

    wb = Workbook()
    ws_account = wb.active
    assert ws_account is not None  # noqa: S101 - openpyxl always returns an active sheet
    ws_account.title = "Konto"
    acc_headers = ["Feld", "Wert"]
    _header_row(ws_account, acc_headers)
    _append(ws_account, ["E-Mail (Anfrage)", email])
    if principal is not None:
        _append(ws_account, ["Login-Subjekt (sub)", principal.get("sub") or ""])
        _append(ws_account, ["E-Mail", principal.get("email") or ""])
        _append(ws_account, ["Anzeigename", principal.get("displayName") or ""])
        _append(ws_account, ["Aktiv", "ja" if principal.get("active") else "nein"])
        _append(ws_account, ["Letzter Login", _fmt_dt(principal.get("lastLogin"))])
    _autosize(ws_account, acc_headers)

    ws_apps = wb.create_sheet(title="Anträge")
    app_headers = [
        "Antrags-ID", "Antragstyp", "Status", "Erstellt", "Antragsteller", "Daten (JSON)",
    ]
    _header_row(ws_apps, app_headers)
    for a in applications:
        _append(
            ws_apps,
            [
                str(a.get("id") or ""),
                a.get("typeName") or "",
                a.get("status") or "",
                _fmt_dt(a.get("createdAt")),
                a.get("applicantName") or "",
                json.dumps(a.get("data") or {}, ensure_ascii=False, sort_keys=True),
            ],
        )
    _autosize(ws_apps, app_headers)

    ws_versions = wb.create_sheet(title="Versionen")
    v_headers = ["Antrags-ID", "Version", "Geändert von", "Zeitpunkt", "Daten (JSON)"]
    _header_row(ws_versions, v_headers)
    for v in versions:
        _append(
            ws_versions,
            [
                str(v.get("applicationId") or ""),
                v.get("version"),
                v.get("changedBy") or "",
                _fmt_dt(v.get("at")),
                json.dumps(v.get("data") or {}, ensure_ascii=False, sort_keys=True),
            ],
        )
    _autosize(ws_versions, v_headers)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _fmt_dt(value: datetime | None) -> str:
    return value.strftime("%Y-%m-%d %H:%M") if value is not None else ""
