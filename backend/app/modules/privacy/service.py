"""DSGVO/Privacy-Services.

* :class:`PrincipalService` — Principal-Erasure (Art. 17): PII leeren, ``sub`` als
  Pseudonym behalten, Sessions verwerfen.
* :class:`ErasureRequestService` — Löschantrags-Queue (open → executed | rejected),
  dispatcht beim Ausführen auf Anonymisierung (Antrag) bzw. Principal-Erasure.
* :class:`PrivacySettingsService` — globaler Aufbewahrungs-Default (Single-Row).

Benachrichtigungen feuert der Router/Worker best-effort *nach* dem Commit
(notifications/privacy.py), nicht die Services selbst.
"""

from __future__ import annotations

from collections.abc import Mapping, Sequence
from datetime import UTC, datetime
from typing import Any
from uuid import UUID

from sqlalchemy import delete, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.modules.admin.models import ApplicationType
from app.modules.applications.models import (
    Applicant,
    Application,
    Comment,
    SubmissionVersion,
)
from app.modules.applications.service import ApplicationsService
from app.modules.audit.actions import AuditAction
from app.modules.audit.service import record as audit_record
from app.modules.auth.models import AuthSession, Principal
from app.modules.files.models import Attachment
from app.modules.files.service import FilesService
from app.modules.flow.models import State
from app.modules.privacy.models import ErasureRequest, PrivacySettings
from app.shared.errors import ConflictError, NotFoundError, ValidationProblem


def _now() -> datetime:
    return datetime.now(UTC)


def _i18n(value: dict[str, str] | None, locale: str) -> str:
    """Lokalisierten Wert (Fallback de→en) aus einem I18n-Map ziehen."""
    data = value or {}
    return data.get(locale) or data.get("de") or data.get("en") or ""


class PrincipalService:
    """Erasure eines Principals (DSGVO Art. 17)."""

    def __init__(self, session: AsyncSession) -> None:
        self.session = session

    async def erase(
        self, principal_id: UUID, *, actor: str, commit: bool = True
    ) -> Principal:
        """PII leeren (email/display_name/calendar_token/oidc_groups → NULL),
        deaktivieren, Sessions verwerfen. ``sub`` bleibt als Pseudonym (Audit-Kette +
        Keycloak-Verknüpfung); die Keycloak-User-Löschung erfolgt **out-of-band** —
        die Rest-Pseudonymität des ``sub`` hängt daran."""
        principal = await self.session.get(Principal, principal_id)
        if principal is None:
            raise NotFoundError(f"principal {principal_id} not found")
        principal.email = None
        principal.display_name = None
        principal.calendar_token = None
        principal.oidc_groups = None
        principal.active = False
        await self.session.execute(
            delete(AuthSession).where(AuthSession.principal_id == principal_id)
        )
        await audit_record(
            self.session,
            actor=actor,
            action=AuditAction.PRINCIPAL_ERASED,
            target_type="principal",
            target_id=str(principal_id),
        )
        if commit:
            await self.session.commit()
        else:
            await self.session.flush()
        return principal


class PrivacySettingsService:
    """Single-Row-Config (globaler Aufbewahrungs-Default in Monaten)."""

    def __init__(self, session: AsyncSession) -> None:
        self.session = session

    async def get(self) -> PrivacySettings:
        settings = await self.session.get(PrivacySettings, 1)
        if settings is None:
            # Defensive: Seed-Zeile fehlt (frisches Schema ohne Migration-Seed) → anlegen.
            settings = PrivacySettings(id=1)
            self.session.add(settings)
            await self.session.flush()
        return settings

    async def update(self, *, default_retention_months: int) -> PrivacySettings:
        settings = await self.get()
        settings.default_retention_months = default_retention_months
        await self.session.commit()
        await self.session.refresh(settings)
        return settings


class ErasureRequestService:
    """Löschantrags-Queue (DSGVO Art. 17): open → executed | rejected."""

    def __init__(self, session: AsyncSession) -> None:
        self.session = session

    async def create(
        self,
        *,
        subject_type: str,
        application_id: UUID | None = None,
        principal_id: UUID | None = None,
        email: str | None = None,
        requested_by: str | None = None,
    ) -> ErasureRequest:
        if subject_type == "applicant" and application_id is None:
            raise ValidationProblem("applicant erasure requires an application_id")
        if subject_type == "principal" and principal_id is None:
            raise ValidationProblem("principal erasure requires a principal_id")
        # Existenz prüfen, statt die FK erst beim Flush brechen zu lassen (404 statt 500).
        if (
            application_id is not None
            and await self.session.get(Application, application_id) is None
        ):
            raise NotFoundError(f"application {application_id} not found")
        if subject_type == "applicant" and email is None and application_id is not None:
            # Antragsteller-Mail jetzt mitnehmen — nach der Anonymisierung ist sie
            # NULL, würde aber für die Bestätigungs-Mail gebraucht.
            email = await self.session.scalar(
                select(Applicant.email).where(
                    Applicant.application_id == application_id
                )
            )
        request = ErasureRequest(
            subject_type=subject_type,
            application_id=application_id,
            principal_id=principal_id,
            email=email,
            status="open",
            requested_by=requested_by,
        )
        self.session.add(request)
        await self.session.flush()
        await audit_record(
            self.session,
            actor=requested_by or "applicant",
            action=AuditAction.ERASURE_REQUESTED,
            target_type="erasure_request",
            target_id=str(request.id),
            data={"subjectType": subject_type},
        )
        await self.session.commit()
        await self.session.refresh(request)
        return request

    async def list(self, *, status: str | None = None) -> list[ErasureRequest]:
        stmt = select(ErasureRequest).order_by(ErasureRequest.created_at.desc())
        if status is not None:
            stmt = stmt.where(ErasureRequest.status == status)
        return list((await self.session.scalars(stmt)).all())

    async def get(self, request_id: UUID) -> ErasureRequest | None:
        return await self.session.get(ErasureRequest, request_id)

    async def _require_open(self, request_id: UUID) -> ErasureRequest:
        request = await self.get(request_id)
        if request is None:
            raise NotFoundError(f"erasure request {request_id} not found")
        if request.status != "open":
            raise ConflictError(
                f"erasure request already {request.status}", code="erasure_not_open"
            )
        return request

    async def execute(
        self, request_id: UUID, *, actor: str, files: FilesService | None = None
    ) -> ErasureRequest:
        """Antrag ausführen: Antrag anonymisieren bzw. Principal löschen, atomar mit
        dem Statuswechsel. ``files`` (mit Storage) ermöglicht das Entfernen der
        Antrags-Anhänge inkl. Storage-Objekte."""
        request = await self._require_open(request_id)
        if request.subject_type == "applicant" and request.application_id is not None:
            await ApplicationsService(self.session).anonymize(
                request.application_id, files=files, actor=actor, commit=False
            )
            await audit_record(
                self.session,
                actor=actor,
                action=AuditAction.ANONYMIZATION,
                target_type="application",
                target_id=str(request.application_id),
            )
        elif request.subject_type == "principal" and request.principal_id is not None:
            await PrincipalService(self.session).erase(
                request.principal_id, actor=actor, commit=False
            )
        else:
            raise ConflictError(
                "erasure request has no resolvable subject", code="erasure_no_subject"
            )
        request.status = "executed"
        request.handled_by = actor
        request.handled_at = _now()
        await audit_record(
            self.session,
            actor=actor,
            action=AuditAction.ERASURE_EXECUTED,
            target_type="erasure_request",
            target_id=str(request.id),
        )
        await self.session.commit()
        await self.session.refresh(request)
        return request

    async def reject(
        self, request_id: UUID, *, actor: str, reason: str | None = None
    ) -> ErasureRequest:
        request = await self._require_open(request_id)
        request.status = "rejected"
        request.handled_by = actor
        request.handled_at = _now()
        request.reason = reason
        await audit_record(
            self.session,
            actor=actor,
            action=AuditAction.ERASURE_REJECTED,
            target_type="erasure_request",
            target_id=str(request.id),
        )
        await self.session.commit()
        await self.session.refresh(request)
        return request


class AuskunftService:
    """DSGVO-Auskunft (Art. 15): alle zu einer E-Mail gespeicherten PII zusammentragen."""

    def __init__(self, session: AsyncSession) -> None:
        self.session = session

    async def collect(self, email: str, *, locale: str = "de") -> dict[str, Any]:
        """Antragsteller-Datensätze (per Mail), zugehörige Anträge + ``data`` +
        Versionshistorie, vom Subjekt einsehbare Kommentare, Anhang-Metadaten und
        den passenden Principal-Datensatz sammeln. Liefert fertige Reihen für
        :func:`build_auskunft_workbook` (Art. 15: Vollständigkeit)."""
        applicants = list(
            (
                await self.session.scalars(
                    select(Applicant).where(Applicant.email == email)
                )
            ).all()
        )
        applicant_by_app = {a.application_id: a for a in applicants}
        app_ids = list(applicant_by_app.keys())

        applications: list[dict[str, Any]] = []
        versions: list[dict[str, Any]] = []
        comments: list[dict[str, Any]] = []
        attachments: list[dict[str, Any]] = []
        if app_ids:
            apps = list(
                (
                    await self.session.scalars(
                        select(Application).where(Application.id.in_(app_ids))
                    )
                ).all()
            )
            type_ids = {a.type_id for a in apps}
            type_names = {
                tid: name
                for tid, name in (
                    await self.session.execute(
                        select(ApplicationType.id, ApplicationType.name_i18n).where(
                            ApplicationType.id.in_(type_ids)
                        )
                    )
                ).all()
            }
            state_ids = {a.current_state_id for a in apps if a.current_state_id}
            state_labels = (
                {
                    sid: label
                    for sid, label in (
                        await self.session.execute(
                            select(State.id, State.label_i18n).where(
                                State.id.in_(state_ids)
                            )
                        )
                    ).all()
                }
                if state_ids
                else {}
            )
            for a in apps:
                applicant = applicant_by_app.get(a.id)
                applications.append(
                    {
                        "id": a.id,
                        "typeName": _i18n(type_names.get(a.type_id), locale),
                        "status": _i18n(state_labels.get(a.current_state_id), locale),
                        "createdAt": a.created_at,
                        "applicantName": applicant.name if applicant else "",
                        "data": a.data,
                    }
                )
            vrows = list(
                (
                    await self.session.scalars(
                        select(SubmissionVersion)
                        .where(SubmissionVersion.application_id.in_(app_ids))
                        .order_by(
                            SubmissionVersion.application_id, SubmissionVersion.version
                        )
                    )
                ).all()
            )
            versions = [
                {
                    "applicationId": v.application_id,
                    "version": v.version,
                    "changedBy": v.changed_by,
                    "at": v.at,
                    "data": v.data,
                }
                for v in vrows
            ]

            # Kommentare, die das Subjekt betreffen/einsehen kann: eigene
            # (author_kind='applicant') sowie öffentliche (für die/den
            # Antragsteller:in sichtbare) Kommentare. Interne Principal-Notizen
            # bleiben außen vor (nicht für das Subjekt bestimmt).
            crows = list(
                (
                    await self.session.scalars(
                        select(Comment)
                        .where(
                            Comment.application_id.in_(app_ids),
                            or_(
                                Comment.author_kind == "applicant",
                                Comment.visibility == "public",
                            ),
                        )
                        .order_by(Comment.application_id, Comment.at)
                    )
                ).all()
            )
            comments = [
                {
                    "applicationId": c.application_id,
                    "authorKind": c.author_kind,
                    "visibility": c.visibility,
                    "body": c.body,
                    "at": c.at,
                }
                for c in crows
            ]

            # Anhang-Metadaten (Dateinamen können PII sein) — ohne das Binär-Objekt.
            arows = list(
                (
                    await self.session.scalars(
                        select(Attachment)
                        .where(Attachment.application_id.in_(app_ids))
                        .order_by(Attachment.application_id, Attachment.created_at)
                    )
                ).all()
            )
            attachments = [
                {
                    "applicationId": at.application_id,
                    "filename": at.filename,
                    "mime": at.mime,
                    "size": at.size,
                    "createdAt": at.created_at,
                }
                for at in arows
            ]

        principal_row = await self.session.scalar(
            select(Principal).where(Principal.email == email)
        )
        principal = (
            {
                "sub": principal_row.sub,
                "email": principal_row.email,
                "displayName": principal_row.display_name,
                "active": principal_row.active,
                "lastLogin": principal_row.last_login,
            }
            if principal_row is not None
            else None
        )
        return {
            "email": email,
            "applications": applications,
            "versions": versions,
            "comments": comments,
            "attachments": attachments,
            "principal": principal,
        }


def _fmt_dt(value: datetime | None) -> str:
    return value.strftime("%Y-%m-%d %H:%M") if value is not None else ""


def build_auskunft_workbook(
    *,
    email: str,
    applications: Sequence[Mapping[str, Any]],
    versions: Sequence[Mapping[str, Any]],
    principal: Mapping[str, Any] | None,
    comments: Sequence[Mapping[str, Any]] = (),
    attachments: Sequence[Mapping[str, Any]] = (),
) -> bytes:
    """DSGVO-Auskunft (Art. 15) als ``.xlsx``-Bytes — vollständig.

    Baut auf dem geteilten :func:`app.shared.xlsx.build_auskunft_workbook`
    (Konto/Anträge/Versionen) auf und ergänzt zwei Blätter für die
    vom Subjekt einsehbaren **Kommentare** und die **Anhang-Metadaten**
    (Dateinamen können PII sein), die der Erasure-Pfad bereits als PII
    behandelt. Reichert die Mappe nachträglich an, statt das geteilte,
    DB-agnostische Workbook-Modul zu ändern."""
    from io import BytesIO

    from openpyxl import load_workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    from app.shared.xlsx import _safe
    from app.shared.xlsx import build_auskunft_workbook as _build_base

    base = _build_base(
        email=email,
        applications=applications,
        versions=versions,
        principal=principal,
    )
    wb = load_workbook(BytesIO(base))

    def _add_sheet(
        title: str,
        headers: Sequence[str],
        rows: Sequence[Sequence[Any]],
    ) -> None:
        ws = wb.create_sheet(title=title)
        ws.append([_safe(h) for h in headers])
        for cell in ws[1]:
            cell.font = Font(bold=True)
        ws.freeze_panes = "A2"
        for row in rows:
            # Formula-Injection neutralisieren (Comment.body / Attachment.filename
            # sind antragstellerseitig befüllt) — gleicher Schutz wie die geteilten
            # Basis-Blätter via app.shared.xlsx._safe.
            ws.append([_safe(cell) for cell in row])
        widths = [len(str(h)) for h in headers]
        for r in ws.iter_rows(min_row=2, values_only=True):
            for i, cell in enumerate(r):
                widths[i] = max(widths[i], len(str(cell)) if cell is not None else 0)
        for i, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = min(width + 2, 60)

    _add_sheet(
        "Kommentare",
        ["Antrags-ID", "Autor-Art", "Sichtbarkeit", "Zeitpunkt", "Text"],
        [
            [
                str(c.get("applicationId") or ""),
                c.get("authorKind") or "",
                c.get("visibility") or "",
                _fmt_dt(c.get("at")),
                c.get("body") or "",
            ]
            for c in comments
        ],
    )
    _add_sheet(
        "Anhänge",
        ["Antrags-ID", "Dateiname", "Typ", "Größe (Bytes)", "Hochgeladen"],
        [
            [
                str(at.get("applicationId") or ""),
                at.get("filename") or "",
                at.get("mime") or "",
                at.get("size"),
                _fmt_dt(at.get("createdAt")),
            ]
            for at in attachments
        ],
    )

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
