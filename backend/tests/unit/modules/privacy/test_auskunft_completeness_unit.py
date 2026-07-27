"""Unit tests for the completeness of the DSGVO Auskunft (Art. 15, AUD-015).

The tests need no database. They prove that `AuskunftService.collect` gathers more
than the applications and the versions. It also collects the comments the subject may
read (own and public) and the attachment metadata. A filename can hold PII. The
module-local `build_auskunft_workbook` then renders two extra sheets from that data.
"""

from __future__ import annotations

from datetime import UTC, datetime
from uuid import uuid4

from app.modules.applications.models import (
    Applicant,
    Application,
    Comment,
    SubmissionVersion,
)
from app.modules.files.models import Attachment
from app.modules.privacy.service import AuskunftService, build_auskunft_workbook
from tests._support.privacy_fakes import fake_session, result

_AT = datetime(2026, 6, 21, 9, 30, 0, tzinfo=UTC)


async def test_collect_gathers_comments_and_attachments() -> None:
    app_id = uuid4()
    type_id = uuid4()
    applicant = Applicant(application_id=app_id, email="z@example.org", name="Z Person")
    application = Application(
        id=app_id, type_id=type_id, current_state_id=None, created_at=_AT, data={}
    )
    version = SubmissionVersion(
        application_id=app_id, version=1, changed_by="applicant", at=_AT, data={}
    )
    comment = Comment(
        application_id=app_id,
        author="z@example.org",
        author_kind="applicant",
        body="Bitte um Rückmeldung",
        visibility="public",
        at=_AT,
    )
    attachment = Attachment(
        application_id=app_id,
        filename="lebenslauf_max_mustermann.pdf",
        mime="application/pdf",
        size=2048,
        storage_key="k",
        created_at=_AT,
    )
    db = fake_session(
        scalars=[
            result(applicant),
            result(application),
            result(version),
            result(comment),
            result(attachment),
        ],
        execute=[result((type_id, {"de": "Typ"}))],  # type_names, no state
        scalar=[None],  # no principal
    )

    data = await AuskunftService(db).collect("z@example.org", locale="de")

    assert len(data["comments"]) == 1
    assert data["comments"][0]["body"] == "Bitte um Rückmeldung"
    assert data["comments"][0]["authorKind"] == "applicant"
    assert data["comments"][0]["visibility"] == "public"

    assert len(data["attachments"]) == 1
    assert data["attachments"][0]["filename"] == "lebenslauf_max_mustermann.pdf"
    assert data["attachments"][0]["mime"] == "application/pdf"
    assert data["attachments"][0]["size"] == 2048


async def test_collect_no_match_has_empty_comment_attachment_lists() -> None:
    db = fake_session(scalars=[result()], scalar=[None])
    data = await AuskunftService(db).collect("nobody@example.org")
    assert data["comments"] == []
    assert data["attachments"] == []


def test_build_workbook_renders_comment_and_attachment_sheets() -> None:
    from io import BytesIO

    from openpyxl import load_workbook

    app_id = uuid4()
    blob = build_auskunft_workbook(
        email="z@example.org",
        applications=[],
        versions=[],
        principal=None,
        comments=[
            {
                "applicationId": app_id,
                "authorKind": "applicant",
                "visibility": "public",
                "at": _AT,
                "body": "ein PII-haltiger Text",
            }
        ],
        attachments=[
            {
                "applicationId": app_id,
                "filename": "rechnung_2026.pdf",
                "mime": "application/pdf",
                "size": 4096,
                "createdAt": _AT,
            }
        ],
    )
    assert isinstance(blob, bytes) and len(blob) > 0

    wb = load_workbook(BytesIO(blob))
    assert "Kommentare" in wb.sheetnames
    assert "Anhänge" in wb.sheetnames

    comment_values = [
        cell for row in wb["Kommentare"].iter_rows(values_only=True) for cell in row
    ]
    assert "ein PII-haltiger Text" in comment_values

    attachment_values = [
        cell for row in wb["Anhänge"].iter_rows(values_only=True) for cell in row
    ]
    assert "rechnung_2026.pdf" in attachment_values


def test_build_workbook_backward_compatible_without_new_kwargs() -> None:
    """The comments and attachments defaults keep old callers with four kwargs valid."""
    from io import BytesIO

    from openpyxl import load_workbook

    blob = build_auskunft_workbook(
        email="z@example.org",
        applications=[],
        versions=[],
        principal=None,
    )
    wb = load_workbook(BytesIO(blob))
    # empty but valid extra sheets
    assert "Kommentare" in wb.sheetnames
    assert "Anhänge" in wb.sheetnames
