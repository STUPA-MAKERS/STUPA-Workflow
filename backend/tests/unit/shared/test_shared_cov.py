"""Zusatz-Coverage für app/shared: xlsx-Workbook-Builder, Guard-Helfer/-Validierung
und Config-Schema-ReDoS-Schutz. Reine Unit-Tests (kein DB/Docker/Netz)."""

from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from typing import Any
from uuid import uuid4

import pytest
from openpyxl import load_workbook
from pydantic import ValidationError

from app.modules.applications.schemas import ApplicationListItem, StateOut
from app.modules.budget.tree_schemas import AllocationView, BudgetTreeNodeOut
from app.shared import xlsx
from app.shared.config_schemas import FieldValidation, _redos_prone
from app.shared.guards import (
    GuardContext,
    GuardError,
    _apply_ordered,
    _to_bool,
    _to_date,
    _to_decimal,
    eval_guard,
    guard_requires_applicant,
    ops_for_type,
    validate_action,
    validate_guard,
)


# --------------------------------------------------------------------------- #
# Hilfen
# --------------------------------------------------------------------------- #
def _load(data: bytes) -> Any:
    """xlsx-Bytes als openpyxl-Workbook laden (für Assertions über Inhalt)."""
    return load_workbook(BytesIO(data))


def _alloc(fy: Any, **kw: Any) -> AllocationView:
    base: dict[str, Any] = dict(
        fiscal_year_id=fy,
        allocated=Decimal("1000"),
        committed=Decimal("200"),
        requested=Decimal("50"),
        available=Decimal("750"),
    )
    base.update(kw)
    return AllocationView.model_validate(base)


def _node(
    name: str,
    *,
    allocs: list[AllocationView] | None = None,
    children: list[BudgetTreeNodeOut] | None = None,
    path_key: str | None = None,
    currency: str = "EUR",
) -> BudgetTreeNodeOut:
    return BudgetTreeNodeOut.model_validate(
        {
            "id": uuid4(),
            "parentId": None,
            "gremiumId": None,
            "key": name,
            "pathKey": path_key or f"VS-{name}",
            "name": name,
            "currency": currency,
            "active": True,
            "byFiscalYear": allocs or [],
            "children": children or [],
        }
    )


def _state(label: dict[str, str] | None) -> StateOut:
    return StateOut.model_validate(
        {
            "id": uuid4(),
            "key": "review",
            "label": label if label is not None else {},
            "editAllowed": True,
        }
    )


def _list_item(**kw: Any) -> ApplicationListItem:
    base: dict[str, Any] = dict(
        id=uuid4(),
        typeId=uuid4(),
        title="Antrag",
        state=None,
        gremiumId=None,
        amount=Decimal("12.50"),
        currency="EUR",
        createdAt=datetime(2026, 6, 1, 9, 30),
        updatedAt=datetime(2026, 6, 2, 10, 0),
    )
    base.update(kw)
    return ApplicationListItem.model_validate(base)


# --------------------------------------------------------------------------- #
# xlsx — Budget-Workbook
# --------------------------------------------------------------------------- #
def test_budget_workbook_empty_tree() -> None:
    """Kein HHJ → ein einzelnes »Budget«-Blatt nur mit Header."""
    data = xlsx.build_budget_workbook([], fiscal_year_labels={})
    wb = _load(data)
    assert wb.sheetnames == ["Budget"]
    ws = wb["Budget"]
    # Header-Zeile vorhanden, keine Datenzeilen.
    assert ws.max_row == 1
    assert ws.cell(row=1, column=1).value == "Kostenstelle"
    assert ws.cell(row=1, column=1).font.bold is True
    assert ws.freeze_panes == "A2"


def test_budget_workbook_one_year_with_and_without_alloc() -> None:
    """Ein HHJ, ein Knoten mit Alloc + ein Kind ohne Alloc (None-Zellen)."""
    fy = uuid4()
    child = _node("child", path_key="VS-root-child")  # keine Alloc → None-Spalten
    root = _node("root", allocs=[_alloc(fy)], children=[child])
    labels = {fy: "Haushalt 2026"}
    data = xlsx.build_budget_workbook([root], fiscal_year_labels=labels)
    wb = _load(data)
    assert wb.sheetnames == ["Haushalt 2026"]
    ws = wb["Haushalt 2026"]
    # Header + root + child = 3 Zeilen.
    assert ws.max_row == 3
    # root: Alloc-Zahlen als float.
    assert ws.cell(row=2, column=1).value == "root"
    assert ws.cell(row=2, column=3).value == 1000.0
    assert ws.cell(row=2, column=5).value == 50.0
    # child: eingerückt + None-Spalten.
    assert ws.cell(row=3, column=1).value == "    child"
    assert ws.cell(row=3, column=3).value is None
    assert ws.cell(row=3, column=7).value == "EUR"


def test_budget_workbook_filter_single_fiscal_year() -> None:
    """``fiscal_year_id`` blendet andere HHJ-Blätter aus."""
    fy1, fy2 = uuid4(), uuid4()
    root = _node("root", allocs=[_alloc(fy1), _alloc(fy2)])
    labels = {fy1: "2025", fy2: "2026"}
    data = xlsx.build_budget_workbook(
        [root], fiscal_year_labels=labels, fiscal_year_id=fy2
    )
    wb = _load(data)
    assert wb.sheetnames == ["2026"]


def test_budget_workbook_repeated_fiscal_year_dedup() -> None:
    """Mehrere Knoten mit demselben HHJ → fy_order dedupliziert (ein Blatt)."""
    fy = uuid4()
    a = _node("a", allocs=[_alloc(fy)])
    b = _node("b", allocs=[_alloc(fy)])  # gleiche fy → kein zweites fy_order-Element
    data = xlsx.build_budget_workbook([a, b], fiscal_year_labels={fy: "2026"})
    wb = _load(data)
    assert wb.sheetnames == ["2026"]
    ws = wb["2026"]
    # Header + a + b = 3 Zeilen, beide Knoten im selben Blatt.
    assert ws.max_row == 3


def test_budget_workbook_missing_label_falls_back_to_hhj() -> None:
    """Fehlendes Label → Default »HHJ« als Blattname."""
    fy = uuid4()
    root = _node("root", allocs=[_alloc(fy)])
    data = xlsx.build_budget_workbook([root], fiscal_year_labels={})
    wb = _load(data)
    assert wb.sheetnames == ["HHJ"]


def test_budget_workbook_multiple_years_stable_order() -> None:
    """Mehrere HHJ → ein Blatt je HHJ, Reihenfolge stabil aus dem Baum."""
    fy1, fy2 = uuid4(), uuid4()
    root = _node("root", allocs=[_alloc(fy1), _alloc(fy2)])
    labels = {fy1: "Erstes", fy2: "Zweites"}
    data = xlsx.build_budget_workbook([root], fiscal_year_labels=labels)
    wb = _load(data)
    assert wb.sheetnames == ["Erstes", "Zweites"]


# --------------------------------------------------------------------------- #
# xlsx — _sheet_title (Sanitisierung + Eindeutigkeit)
# --------------------------------------------------------------------------- #
def test_sheet_title_sanitizes_illegal_chars() -> None:
    used: set[str] = set()
    title = xlsx._sheet_title("A[b]:c*d?e/f\\g", used)
    assert title == "A_b__c_d_e_f_g"
    assert "[" not in title and ":" not in title


def test_sheet_title_empty_becomes_hhj() -> None:
    used: set[str] = set()
    # Nur illegale Zeichen + Whitespace → nach strip() leer → »HHJ«.
    assert xlsx._sheet_title("   ", used) == "HHJ"


def test_sheet_title_truncates_to_31() -> None:
    used: set[str] = set()
    long = "x" * 50
    title = xlsx._sheet_title(long, used)
    assert len(title) == 31


def test_sheet_title_dedupes_with_suffix() -> None:
    used: set[str] = set()
    a = xlsx._sheet_title("Haushalt", used)
    b = xlsx._sheet_title("Haushalt", used)
    c = xlsx._sheet_title("Haushalt", used)
    assert a == "Haushalt"
    assert b == "Haushalt (2)"
    assert c == "Haushalt (3)"


def test_budget_workbook_duplicate_labels_get_unique_sheets() -> None:
    fy1, fy2 = uuid4(), uuid4()
    root = _node("root", allocs=[_alloc(fy1), _alloc(fy2)])
    labels = {fy1: "Topf", fy2: "Topf"}
    data = xlsx.build_budget_workbook([root], fiscal_year_labels=labels)
    wb = _load(data)
    assert wb.sheetnames == ["Topf", "Topf (2)"]


# --------------------------------------------------------------------------- #
# xlsx — Antragsliste
# --------------------------------------------------------------------------- #
def test_applications_workbook_full_row() -> None:
    item = _list_item(
        state=_state({"de": "Prüfung", "en": "Review"}),
        gremiumId=uuid4(),
    )
    type_names = {item.type_id: "Finanzantrag"}
    gremium_names = {item.gremium_id: "StuPa"}
    data = xlsx.build_applications_workbook(
        [item], type_names=type_names, gremium_names=gremium_names
    )
    wb = _load(data)
    assert wb.sheetnames == ["Anträge"]
    ws = wb["Anträge"]
    assert ws.cell(row=2, column=1).value == "Antrag"
    assert ws.cell(row=2, column=2).value == "Finanzantrag"
    assert ws.cell(row=2, column=3).value == "Prüfung"
    assert ws.cell(row=2, column=4).value == "StuPa"
    assert ws.cell(row=2, column=5).value == 12.5
    assert ws.cell(row=2, column=7).value == "2026-06-01 09:30"


def test_applications_workbook_locale_label_fallbacks() -> None:
    # locale=fr nicht vorhanden → de → en Fallback-Kette.
    only_en = _list_item(state=_state({"en": "Approved"}))
    data = xlsx.build_applications_workbook(
        [only_en], type_names={}, gremium_names={}, locale="fr"
    )
    ws = _load(data)["Anträge"]
    assert ws.cell(row=2, column=3).value == "Approved"


def test_applications_workbook_empty_label_dict() -> None:
    # state vorhanden, aber leeres Label → leerer String (keine der Locales matched).
    blank = _list_item(state=_state({}))
    data = xlsx.build_applications_workbook(
        [blank], type_names={}, gremium_names={}
    )
    ws = _load(data)["Anträge"]
    assert ws.cell(row=2, column=3).value is None or ws.cell(row=2, column=3).value == ""


def test_applications_workbook_no_state_no_gremium_no_title() -> None:
    item = _list_item(state=None, gremiumId=None, title=None, amount=None, currency=None)
    data = xlsx.build_applications_workbook(
        [item], type_names={}, gremium_names={}
    )
    ws = _load(data)["Anträge"]
    # title None → "", gremium leer → "", state None → "".
    assert ws.cell(row=2, column=1).value is None or ws.cell(row=2, column=1).value == ""
    assert ws.cell(row=2, column=4).value is None or ws.cell(row=2, column=4).value == ""
    assert ws.cell(row=2, column=5).value is None  # _num(None)


# --------------------------------------------------------------------------- #
# xlsx — Buchungen
# --------------------------------------------------------------------------- #
class _Expense:
    def __init__(self, **kw: Any) -> None:
        self.created_at = kw.get("created_at", datetime(2026, 1, 5, 8, 0))
        self.kind = kw.get("kind", "expense")
        self.description = kw.get("description", "Druckkosten")
        self.path_key = kw.get("path_key", "VS-800-40")
        self.application_title = kw.get("application_title", "Plakate")
        self.account_name = kw.get("account_name", "Konto A")
        self.amount = kw.get("amount", Decimal("42.00"))
        self.currency = kw.get("currency", "EUR")


def test_expenses_workbook_de_labels() -> None:
    items = [_Expense(kind="expense"), _Expense(kind="income")]
    data = xlsx.build_expenses_workbook(items)
    ws = _load(data)["Buchungen"]
    assert ws.cell(row=2, column=2).value == "Ausgabe"
    assert ws.cell(row=3, column=2).value == "Einnahme"
    assert ws.cell(row=2, column=7).value == 42.0


def test_expenses_workbook_en_labels_and_unknown_kind() -> None:
    items = [_Expense(kind="income"), _Expense(kind="weird")]
    data = xlsx.build_expenses_workbook(items, locale="en")
    ws = _load(data)["Buchungen"]
    assert ws.cell(row=2, column=2).value == "Income"
    # Unbekannte Art → fällt auf den Roh-Wert zurück.
    assert ws.cell(row=3, column=2).value == "weird"


def test_expenses_workbook_none_fields() -> None:
    items = [
        _Expense(
            description=None,
            path_key=None,
            application_title=None,
            account_name=None,
            amount=None,
            currency=None,
        )
    ]
    data = xlsx.build_expenses_workbook(items)
    ws = _load(data)["Buchungen"]
    assert ws.cell(row=2, column=3).value is None or ws.cell(row=2, column=3).value == ""
    assert ws.cell(row=2, column=7).value is None  # _num(None)


# --------------------------------------------------------------------------- #
# xlsx — DSGVO-Auskunft
# --------------------------------------------------------------------------- #
def test_auskunft_workbook_with_principal_and_rows() -> None:
    principal = {
        "sub": "auth0|abc",
        "email": "a@b.c",
        "displayName": "Alice",
        "active": True,
        "lastLogin": datetime(2026, 6, 1, 12, 0),
    }
    applications = [
        {
            "id": uuid4(),
            "typeName": "Finanzantrag",
            "status": "review",
            "createdAt": datetime(2026, 1, 1, 0, 0),
            "applicantName": "Alice",
            "data": {"b": 2, "a": 1},
        }
    ]
    versions = [
        {
            "applicationId": uuid4(),
            "version": 1,
            "changedBy": "Alice",
            "at": datetime(2026, 1, 2, 0, 0),
            "data": {"x": 1},
        }
    ]
    data = xlsx.build_auskunft_workbook(
        email="a@b.c",
        applications=applications,
        versions=versions,
        principal=principal,
    )
    wb = _load(data)
    assert wb.sheetnames == ["Konto", "Anträge", "Versionen"]
    konto = wb["Konto"]
    assert konto.cell(row=2, column=2).value == "a@b.c"  # E-Mail (Anfrage)
    assert konto.cell(row=3, column=2).value == "auth0|abc"
    assert konto.cell(row=6, column=2).value == "ja"  # Aktiv=True
    apps = wb["Anträge"]
    # JSON sortiert (sort_keys) → "a" vor "b".
    assert apps.cell(row=2, column=6).value == '{"a": 1, "b": 2}'
    versions_ws = wb["Versionen"]
    assert versions_ws.cell(row=2, column=2).value == 1


def test_auskunft_workbook_inactive_principal() -> None:
    principal = {
        "sub": None,
        "email": None,
        "displayName": None,
        "active": False,
        "lastLogin": None,
    }
    data = xlsx.build_auskunft_workbook(
        email="x@y.z", applications=[], versions=[], principal=principal
    )
    konto = _load(data)["Konto"]
    assert konto.cell(row=6, column=2).value == "nein"  # active False
    # Letzter Login None → "" via _fmt_dt.
    assert konto.cell(row=7, column=2).value is None or konto.cell(row=7, column=2).value == ""


def test_auskunft_workbook_no_principal() -> None:
    data = xlsx.build_auskunft_workbook(
        email="x@y.z", applications=[], versions=[], principal=None
    )
    wb = _load(data)
    konto = wb["Konto"]
    # Nur Header + E-Mail-Zeile, keine Principal-Zeilen.
    assert konto.max_row == 2


def test_auskunft_workbook_empty_app_and_version_data() -> None:
    # data fehlt → json.dumps({}) ; ids fehlen → leerer String.
    applications = [{"typeName": None, "status": None, "applicantName": None}]
    versions = [{"version": None, "changedBy": None}]
    data = xlsx.build_auskunft_workbook(
        email="x@y.z", applications=applications, versions=versions, principal=None
    )
    wb = _load(data)
    assert wb["Anträge"].cell(row=2, column=6).value == "{}"
    assert wb["Versionen"].cell(row=2, column=5).value == "{}"


# --------------------------------------------------------------------------- #
# xlsx — Klein-Helfer
# --------------------------------------------------------------------------- #
def test_num_and_fmt_dt_helpers() -> None:
    assert xlsx._num(None) is None
    assert xlsx._num(Decimal("3.5")) == 3.5
    assert xlsx._num(7) == 7.0
    assert xlsx._fmt_dt(None) == ""
    assert xlsx._fmt_dt(datetime(2026, 6, 16, 14, 5)) == "2026-06-16 14:05"


def test_autosize_caps_width_at_60() -> None:
    # Eine sehr lange Zelle → Spaltenbreite auf 60 gedeckelt.
    long_name = "z" * 200
    fy = uuid4()
    root = _node(long_name, allocs=[_alloc(fy)])
    data = xlsx.build_budget_workbook([root], fiscal_year_labels={fy: "HHJ"})
    ws = _load(data)["HHJ"]
    assert ws.column_dimensions["A"].width == 60


# --------------------------------------------------------------------------- #
# guards — Koerzierungshelfer
# --------------------------------------------------------------------------- #
def test_to_decimal_variants() -> None:
    assert _to_decimal(None) is None
    assert _to_decimal(True) is None  # bool ausgeschlossen
    assert _to_decimal(False) is None
    assert _to_decimal("3.14") == Decimal("3.14")
    assert _to_decimal(5) == Decimal("5")
    assert _to_decimal("notanumber") is None  # InvalidOperation
    assert _to_decimal(object()) is None  # TypeError → None


def test_to_date_variants() -> None:
    assert _to_date(datetime(2026, 6, 16, 1, 2)) == date(2026, 6, 16)
    assert _to_date(date(2026, 6, 16)) == date(2026, 6, 16)
    assert _to_date("2026-06-16T10:00:00") == date(2026, 6, 16)
    assert _to_date("not-a-date") is None  # ValueError → None
    assert _to_date(123) is None  # nicht-str/date/datetime → None


def test_to_bool_variants() -> None:
    assert _to_bool(True) is True
    assert _to_bool(False) is False
    assert _to_bool("Ja") is True
    assert _to_bool("on") is True
    assert _to_bool("nope") is False
    assert _to_bool(1) is True
    assert _to_bool(0) is False
    assert _to_bool([1]) is True  # bool(value) Fallback


def test_ops_for_type() -> None:
    assert "<" in ops_for_type("number")
    assert "<" in ops_for_type("currency")
    assert "<" in ops_for_type("date")
    assert ops_for_type("bool") == frozenset({"=="})
    assert "in" in ops_for_type("text")
    assert "in" in ops_for_type("anything_else")  # Default → Text-Ops


def test_apply_ordered_unknown_op_is_false() -> None:
    assert _apply_ordered("==", 1, 1) is True
    assert _apply_ordered("<", 1, 2) is True
    assert _apply_ordered("~~", 1, 2) is False  # unbekannt → False


# --------------------------------------------------------------------------- #
# guards — eval_guard compare edge cases
# --------------------------------------------------------------------------- #
def test_compare_requires_object() -> None:
    with pytest.raises(GuardError, match="object"):
        eval_guard({"compare": "notdict"}, GuardContext())


def test_compare_field_must_be_nonempty_string() -> None:
    with pytest.raises(GuardError, match="non-empty string"):
        eval_guard({"compare": {"field": "", "op": "=="}}, GuardContext())
    with pytest.raises(GuardError, match="non-empty string"):
        eval_guard({"compare": {"field": 7, "op": "=="}}, GuardContext())


def test_compare_op_must_be_string() -> None:
    with pytest.raises(GuardError, match="op must be a string"):
        eval_guard({"compare": {"field": "x", "op": 5}}, GuardContext())


def test_compare_empty_left_value_is_false() -> None:
    ctx = GuardContext(field_values={"x": ""}, field_types={"x": "text"})
    assert eval_guard({"compare": {"field": "x", "op": "==", "value": ""}}, ctx) is False
    ctx_none = GuardContext(field_values={"x": None}, field_types={"x": "text"})
    assert eval_guard({"compare": {"field": "x", "op": "==", "value": ""}}, ctx_none) is False


def test_compare_numeric_uncoercible_operand_is_false() -> None:
    ctx = GuardContext(field_values={"amount": "5"}, field_types={"amount": "number"})
    # Operand nicht in Decimal koerzierbar → False.
    assert eval_guard(
        {"compare": {"field": "amount", "op": ">", "value": "x"}}, ctx
    ) is False


def test_compare_date_uncoercible_is_false() -> None:
    ctx = GuardContext(field_values={"due": "2026-06-01"}, field_types={"due": "date"})
    assert eval_guard(
        {"compare": {"field": "due", "op": ">", "value": "not-a-date"}}, ctx
    ) is False


def test_compare_text_in_non_list_is_false() -> None:
    ctx = GuardContext(field_values={"k": "a"}, field_types={"k": "text"})
    # op "in" mit Nicht-Liste → False (kein Crash).
    assert eval_guard({"compare": {"field": "k", "op": "in", "value": "abc"}}, ctx) is False


def test_compare_text_unknown_op_after_type_check_is_false() -> None:
    # Für text ist nur ==,!=,in erlaubt; ein numerischer Operator wird vorher per
    # ops_for_type ausgesiebt (fail-closed False).
    ctx = GuardContext(field_values={"k": "a"}, field_types={"k": "text"})
    assert eval_guard({"compare": {"field": "k", "op": "<", "value": "b"}}, ctx) is False


# --------------------------------------------------------------------------- #
# guards — guard_requires_applicant
# --------------------------------------------------------------------------- #
def test_guard_requires_applicant_direct() -> None:
    assert guard_requires_applicant({"actorIsApplicant": True}) is True


def test_guard_requires_applicant_nested_in_combinator() -> None:
    g = {"and": [{"roleIs": "x"}, {"or": [{"actorIsApplicant": True}]}]}
    assert guard_requires_applicant(g) is True


def test_guard_requires_applicant_combinator_single_child_not_list() -> None:
    # 'not' mit einem einzelnen (Nicht-Listen-)Kind, das das Gate enthält.
    assert guard_requires_applicant({"not": {"actorIsApplicant": True}}) is True


def test_guard_requires_applicant_false_cases() -> None:
    assert guard_requires_applicant(None) is False
    assert guard_requires_applicant({}) is False  # len != 1
    assert guard_requires_applicant({"a": 1, "b": 2}) is False  # mehrere Keys
    assert guard_requires_applicant("notadict") is False  # type: ignore[arg-type]
    assert guard_requires_applicant({"roleIs": "x"}) is False  # Leaf ohne Gate
    # Kombinator ohne das Gate + Nicht-dict-Kinder werden ignoriert.
    assert guard_requires_applicant({"and": [{"roleIs": "x"}, "junk"]}) is False


# --------------------------------------------------------------------------- #
# guards — validate_guard / _validate_compare zusätzliche Zweige
# --------------------------------------------------------------------------- #
def test_validate_guard_multi_key() -> None:
    with pytest.raises(GuardError, match="exactly one operator"):
        validate_guard({"roleIs": "x", "hasField": "y"})


def test_validate_compare_not_an_object() -> None:
    with pytest.raises(GuardError, match="object"):
        validate_guard({"compare": "notdict"})


def test_validate_compare_field_must_be_string() -> None:
    with pytest.raises(GuardError, match="compare.field"):
        validate_guard({"compare": {"field": 1, "op": "=="}})
    with pytest.raises(GuardError, match="compare.field"):
        validate_guard({"compare": {"field": "", "op": "=="}})


def test_validate_compare_in_requires_list_ok() -> None:
    validate_guard({"compare": {"field": "x", "op": "in", "value": ["a", "b"]}})


# --------------------------------------------------------------------------- #
# guards — validate_action notify zusätzliche Zweige
# --------------------------------------------------------------------------- #
def test_validate_action_notify_recipient_not_dict() -> None:
    with pytest.raises(GuardError, match="must be an object"):
        validate_action({"type": "notify", "recipients": ["junk"]})


def test_validate_action_notify_applicant_must_not_have_ref() -> None:
    with pytest.raises(GuardError, match="must not have 'ref'"):
        validate_action(
            {"type": "notify", "recipients": [{"kind": "applicant", "ref": "x"}]}
        )


def test_validate_action_notify_role_and_email_require_ref() -> None:
    with pytest.raises(GuardError, match="requires 'ref'"):
        validate_action({"type": "notify", "recipients": [{"kind": "role"}]})
    with pytest.raises(GuardError, match="requires 'ref'"):
        validate_action({"type": "notify", "recipients": [{"kind": "email"}]})


def test_validate_action_notify_recipients_not_a_list() -> None:
    with pytest.raises(GuardError, match="non-empty 'recipients' list"):
        validate_action({"type": "notify", "recipients": "all"})


# --------------------------------------------------------------------------- #
# config_schemas — _redos_prone + FieldValidation.pattern
# --------------------------------------------------------------------------- #
def test_redos_prone_detects_nested_unbounded() -> None:
    # Klassische katastrophale Form: Repeat innerhalb eines wiederholten Subpatterns.
    assert _redos_prone(r"(a+)+") is True
    assert _redos_prone(r"(a*)*") is True
    assert _redos_prone(r"([ab]+)+") is True


def test_redos_prone_allows_safe_patterns() -> None:
    assert _redos_prone(r"^[A-Z]{2}\d{2}$") is False
    assert _redos_prone(r"abc") is False
    assert _redos_prone(r"a+") is False  # nur ein einfacher Quantor
    assert _redos_prone(r"[0-9]*") is False


def test_redos_prone_branch_inside_repeat() -> None:
    # BRANCH-Kinder werden traversiert; ein verschachtelter unbegrenzter Quantor
    # innerhalb eines wiederholten Branch wird erkannt.
    assert _redos_prone(r"(a+|b)+") is True


def test_redos_prone_repeat_nested_in_repeat_body() -> None:
    # ``has_unbounded_repeat`` muss durch eine MAX_REPEAT in der Body-Traversierung
    # rekursieren (children() für MAX_REPEAT → av[2]).
    assert _redos_prone(r"(a+b+)+") is True


def test_redos_prone_has_unbounded_recurses_through_bounded_repeat() -> None:
    # Äußeres unbegrenztes ``+`` → has_unbounded_repeat(body); der Body enthält ein
    # BEGRENZTES ``{2}`` (kein maxrepeat-Treffer), das children() für MAX_REPEAT
    # durchläuft und so das innere ``a+`` findet.
    assert _redos_prone(r"((a+){2})+") is True


def test_redos_prone_bounded_outer_with_nested_inner() -> None:
    # Äußerer Quantor begrenzt ({1,2}); ``walk(body)`` rekursiert und findet das
    # innere nested-unbounded (a+)+ → True.
    assert _redos_prone(r"((a+)+){1,2}") is True


def test_redos_prone_top_level_branch_with_redos_alternative() -> None:
    # Top-Level-BRANCH ohne umschließenden Repeat: walk's elif-Zweig traversiert die
    # Branch-Kinder, eines davon ist redos-anfällig.
    assert _redos_prone(r"(a+)+|safe") is True


def test_redos_prone_bounded_outer_repeat_is_safe() -> None:
    # Äußerer Quantor begrenzt ({1,3}) → has_unbounded_repeat-Pfad greift nicht,
    # walk geht in den Body weiter (kein nested-unbounded-Treffer).
    assert _redos_prone(r"(a+){1,3}") is False


def test_redos_prone_invalid_pattern_returns_false() -> None:
    # Unparsebares Pattern → walk wirft, äußeres except → False (nur Längen-Schranke).
    assert _redos_prone(r"(") is False


def test_field_validation_pattern_ok() -> None:
    v = FieldValidation.model_validate({"pattern": r"^\d{4}$"})
    assert v.pattern == r"^\d{4}$"


def test_field_validation_pattern_none() -> None:
    # Explizites ``pattern=None`` lässt den Validator laufen (None-Frühausstieg).
    v = FieldValidation.model_validate({"min": 1, "pattern": None})
    assert v.pattern is None


def test_field_validation_pattern_too_long_rejected() -> None:
    with pytest.raises(ValidationError, match="too long"):
        FieldValidation.model_validate({"pattern": "a" * 201})


def test_field_validation_pattern_redos_rejected() -> None:
    with pytest.raises(ValidationError, match="ReDoS"):
        FieldValidation.model_validate({"pattern": r"(a+)+"})
